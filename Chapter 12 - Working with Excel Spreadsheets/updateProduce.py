#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Sep  8 15:38:53 2019

@author: nishanth
"""

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')

wb.sheetnames

sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for rowNum in (2, sheet.max_row):
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName]
        
        
wb.save('updatedProduceSales.xlsx')