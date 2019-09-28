#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep  9 19:20:28 2019

@author: nishanth
"""

import openpyxl, sys
from openpyxl.utils import get_column_letter, column_index_from_string

if len(sys.argv) == 1:
    print("Please provide arguments")
else:
    x = int(sys.argv)
    
wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb['Sheet']

c = sheet.cell(row = x + 1, column = x+1 )

c.column
c.coordinate

get_column_letter(x+1)
column_index_from_string(sheet['A1'].column)



for i in range(x+1):
    sheet.cell(row = 1, column = i+1).value = i
    sheet.cell(row = i+1 , column = 1).value = i
    
for rowOfCellObjects in sheet['B2':str(c.coordinate)]:
    for cellObject in rowOfCellObjects:
        cellObject.value = sheet.cell(row = cellObject.row, column = 1).value * sheet.cell(row = 1 ,column = column_index_from_string(str(cellObject.column))).value
        print(cellObject.value)
    
wb.save('multiplication.xlsx')

