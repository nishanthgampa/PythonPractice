#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep  4 20:12:00 2019

@author: nishanth
"""

import openpyxl,os, datetime

pwd()
os.chdir('/Users/nishanth/Desktop/GIT/PythonPractice')
os.chdir('Chapter 12 - Working with Excel Spreadsheets')

wb = openpyxl.load_workbook('example.xlsx') #open workbook
type(wb)

wb.get_sheet_names() # gets you all the sheets names

sheet = wb.get_sheet_by_name('Sheet3')

type(sheet)

sheet.title

anotherSheet = wb.get_active_sheet()

anotherSheet

sheet = wb.get_sheet_by_name('Sheet1')

sheet['A1']

sheet['A1'].value

c = sheet['B1']

c.value

'Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + str(c.value)

'Cell ' + c.coordinate + ' is ' + c.value

sheet['C1'].value

sheet.cell(row = 1, column = 2).value

for i in range(1,8,2):
    print(i,sheet.cell(row = i, column=2).value)
    
sheet.max_row

sheet.max_column

from openpyxl.utils import get_column_letter, column_index_from_string

get_column_letter(900)

column_index_from_string('AA')

tuple(sheet['A1':'C3'])

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print("#######End of Row########")


for i in sheet['B']:
    print(i.value)

pwd()

wb1 = openpyxl.Workbook()

wb1.sheetnames

sheet = wb1['Sheet']

sheet.title = 'This sheet is created via Python'

wb1.save('PythonCreatedSheet.xlsx')

wb1.create_sheet()

wb1.sheetnames

wb1.create_sheet(index = 0, title = 'First Sheet')

wb1.sheetnames

wb1.create_sheet(index = 1, title = 'Second Sheet')

wb1.sheetnames

wb1.remove(wb1['This sheet is created via Python'])

sheet = wb1['First Sheet']

sheet['A1'] = 'Hello World!'

sheet['A1'].value





